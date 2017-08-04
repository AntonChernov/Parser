import asyncio
from bs4 import BeautifulSoup
from openpyxl import Workbook
import pandas as pd
import urllib.request
import time
import re


async def create_xlsx_file(data=None, filepath=None):
    wb = Workbook()
    qw = wb.active
    for i in range(1, len(data[0]) + 1):
        qw.cell(row=1, column=i).value = data[0][i - 1]
    for j in range(2, len(data[1]) + 2):
        for val in data[1][j-2]:
            qw.cell(row=j, column=data[1][j-2].index(val) + 1).value = val
    wb.save(filename='test.xlsx')


async def parse_thead(thead=None):
    return [str(i.string) for i in thead.thead.find_all('th') if i.string != '#']


async def parse_tbody(tbody=None):
    return [i.find_all('td') for i in tbody.tbody.find_all('tr')]


async def parse_tbody_td_data(td_lists=None):
    return [
        [
            str(i[1].a.string),
            str(i[2].string),
            float(i[3]['data-usd']) if i[3]['data-usd'] != '?' else float(0),
            float(i[4].find(attrs={'data-usd': re.compile(r"\d*")})['data-usd']) if i[4].a['data-usd'] != 'None' and
                                                                                    i[4].a['data-usd'] != '?' else 0.0,
            float(i[5].find(attrs={'data-supply': re.compile(r".*")})['data-supply']) if i[5].find(
                attrs={'data-supply': re.compile(r".*")})['data-supply'] != 'None' and i[5].find(
                attrs={'data-supply': re.compile(r".*")})['data-supply'] != '?' else 0.0,
            float(i[6].find(attrs={'data-usd': re.compile(r".*")})['data-usd']) if i[6].find(
                attrs={'data-usd': re.compile(r".*")})['data-usd'] != 'None' and i[6].find(
                attrs={'data-usd': re.compile(r".*")})['data-usd'] != '?' else 0.0,
            float(i[7].contents[0][:-1]) if i[7].contents[0][:-1] != '' else 0.0,
            float(i[8].contents[0][:-1]) if i[8].contents[0][:-1] != '' else 0.0,
            float(i[9].contents[0][:-1]) if i[9].contents[0][:-1] != '' else 0.0,
        ] for i in td_lists
    ]


async def parse_html(html=None):
    return BeautifulSoup(html, 'html.parser')


async def parser(url):
    with urllib.request.urlopen(url=url) as pars:
        html = pars.read()

    data = []
    soup = await parse_html(html=html)

    data.append(await parse_thead(thead=soup))
    data.append(await parse_tbody_td_data(td_lists=await parse_tbody(tbody=soup)))
    await create_xlsx_file(data=data)


if __name__ == '__main__':
    a = time.time()
    loop = asyncio.get_event_loop()
    loop.run_until_complete(parser('https://coinmarketcap.com/all/views/all/'))
    loop.close()
    print(time.time() - a)
